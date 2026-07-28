"""Lettura input (professori/classi/materie) e scrittura output (orario) da/verso Excel."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def leggi_input(path):
    """Legge il file Excel di input e restituisce (professori, classi, materie, aule_preferenze).

    professori: dict {nome: {"materie": [...], "max_ore": int,
                              "giorno_preferito": str o None, "aula": str (opzionale)}}
    classi: list di stringhe
    materie: dict {materia: ore_settimanali}
    aule_preferenze: dict {aula: giorno_preferito}  (vuoto se non specificato)

    Colonne opzionali nel foglio "Professori" (in coda, se presenti):
      - GiornoPreferito: giorno della settimana in cui il professore preferisce
        essere impegnato (es. "Lun"). È una preferenza, non un vincolo rigido.
      - Aula: nome dell'aula assegnata al professore. Se assente, viene generata
        automaticamente come "Aula_<Nome>".

    Foglio opzionale "Aule" con colonne Aula, GiornoPreferito: indica il giorno
    in cui si preferisce che quell'aula venga usata.
    """
    try:
        xls = pd.read_excel(path, sheet_name=None)
    except Exception as e:
        raise ValueError(f"Impossibile leggere il file Excel: {e}")

    # Leggi Professori
    if "Professori" not in xls:
        raise ValueError("Foglio 'Professori' non trovato nel file Excel.")

    df_prof = xls["Professori"]
    ha_col_giorno = df_prof.shape[1] > 3
    ha_col_aula = df_prof.shape[1] > 4

    def _valore_opzionale(row, indice):
        if indice >= len(row):
            return None
        val = row.iloc[indice]
        if pd.isna(val):
            return None
        val = str(val).strip()
        return val if val and val.lower() != "nan" else None

    professori = {}
    for idx, row in df_prof.iterrows():
        try:
            nome = str(row.iloc[0]).strip()
            materie_prof = [m.strip() for m in str(row.iloc[1]).split(",") if m.strip()]
            max_ore = int(row.iloc[2])
            professori[nome] = {"materie": materie_prof, "max_ore": max_ore}

            if ha_col_giorno:
                giorno_pref = _valore_opzionale(row, 3)
                if giorno_pref:
                    professori[nome]["giorno_preferito"] = giorno_pref

            if ha_col_aula:
                aula = _valore_opzionale(row, 4)
                if aula:
                    professori[nome]["aula"] = aula
        except Exception as e:
            raise ValueError(f"Errore nella riga {idx+2} del foglio 'Professori': {e}")

    # Leggi Classi
    if "Classi" not in xls:
        raise ValueError("Foglio 'Classi' non trovato nel file Excel.")
    
    df_classi = xls["Classi"]
    classi = []
    for idx, row in df_classi.iterrows():
        try:
            classe = str(row.iloc[0]).strip()
            if classe:
                classi.append(classe)
        except Exception as e:
            raise ValueError(f"Errore nella riga {idx+2} del foglio 'Classi': {e}")

    # Leggi Materie
    if "Materie" not in xls:
        raise ValueError("Foglio 'Materie' non trovato nel file Excel.")
    
    df_materie = xls["Materie"]
    materie = {}
    for idx, row in df_materie.iterrows():
        try:
            materia = str(row.iloc[0]).strip()
            ore = int(row.iloc[1])
            materie[materia] = ore
        except Exception as e:
            raise ValueError(f"Errore nella riga {idx+2} del foglio 'Materie': {e}")

    # Leggi Aule (opzionale)
    aule_preferenze = {}
    if "Aule" in xls:
        df_aule = xls["Aule"]
        for idx, row in df_aule.iterrows():
            try:
                aula = str(row.iloc[0]).strip()
                if not aula or aula.lower() == "nan":
                    continue
                giorno = _valore_opzionale(row, 1)
                if giorno:
                    aule_preferenze[aula] = giorno
            except Exception as e:
                raise ValueError(f"Errore nella riga {idx+2} del foglio 'Aule': {e}")

    return professori, classi, materie, aule_preferenze


def scrivi_output(path, orario, classi, giorni, ore_per_giorno):
    """Scrive l'orario generato in un Excel con un foglio per classe.

    orario: dict {classe: {(giorno, ora): "Materia (Prof)"}}
    giorni: list di stringhe (es. ["Lun", "Mar", "Mer", "Gio", "Ven"])
    ore_per_giorno: int
    """
    wb = Workbook()
    # Rimuovi il foglio predefinito
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="4472C4")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for classe in classi:
        # Crea un foglio per la classe
        sheet_name = classe[:31]  # Excel limita a 31 caratteri
        ws = wb.create_sheet(sheet_name)
        
        # Intestazione: Ora e giorni
        ws.cell(row=1, column=1, value="Ora")
        for j, g in enumerate(giorni, start=2):
            ws.cell(row=1, column=j, value=g)
        
        # Applica stili all'intestazione
        for col in range(1, len(giorni) + 2):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # Riempie le ore
        dati_classe = orario.get(classe, {})
        
        for h in range(ore_per_giorno):
            # Numero dell'ora
            cell_ora = ws.cell(row=h + 2, column=1, value=h + 1)
            cell_ora.alignment = center
            cell_ora.border = border
            
            # Materie per ogni giorno
            for j, g in enumerate(giorni, start=2):
                valore = dati_classe.get((g, h), "")
                cell = ws.cell(row=h + 2, column=j, value=valore)
                cell.alignment = center
                cell.border = border
                if valore:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", start_color="EAF1FB")

        # Imposta larghezze colonne
        ws.column_dimensions["A"].width = 8
        for j in range(2, len(giorni) + 2):
            col_letter = chr(64 + j)  # Converte numero a lettera (2->B, 3->C, ...)
            ws.column_dimensions[col_letter].width = 25
        
        # Imposta altezza righe
        for r in range(2, ore_per_giorno + 2):
            ws.row_dimensions[r].height = 30

    # Salva il file
    try:
        wb.save(path)
    except Exception as e:
        raise ValueError(f"Impossibile salvare il file: {e}")